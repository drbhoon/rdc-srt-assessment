#!/usr/bin/env python3
"""
Re-validation analysis: compare the engine's (AI) competency scores against the
human (BH) reference after re-baselining under skill v2.5.

It reads BH scores from the validation workbook and AI scores from EITHER:
  (a) the live Postgres DB (set DATABASE_URL), matching candidates by employee
      name (case-insensitive), OR
  (b) a fresh export of the same workbook whose "AI Rating" rows have been
      updated post-rescore (pass --sheet-only).

Outputs: bias, MAE, correlation, within ±1 / ±2, per-competency gap, and the
top outliers (|gap| > 2). Read-only — never writes to the DB.

Usage:
    python revalidation_analysis.py                 # DB AI vs sheet BH (needs DATABASE_URL)
    python revalidation_analysis.py --sheet-only     # AI + BH both from the sheet
    python revalidation_analysis.py --xlsx "<path>"  # override workbook path
"""
import argparse
import json
import os
import statistics
import sys
from collections import OrderedDict, defaultdict

import openpyxl

DEFAULT_XLSX = os.path.join(
    os.path.dirname(__file__),
    "SRT based Competency validation - Self-AI-BH May 26.xlsx",
)

# Sheet competency columns J..S (10..19, 1-indexed) → canonical names.
COMP_COLS = list(range(10, 20))
COMP_NAMES = [
    "Operational Discipline & SARTAJ Ownership",
    "Cost & Resource Responsibility",
    "Functional Knowledge & Multiskilling",
    "Team Orientation & Delegation",
    "Communication & Assertiveness",
    "Vendor & External Stakeholder Management",
    "Integrity & Trust",
    "Customer Orientation & Relationship Handling",
    "Preventive Maintenance & Asset Care",
    "Planning, Organizing & Coordination",
]


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace("%", ""))
    except ValueError:
        return None


def load_sheet(path):
    """Return {emp_code: {'name':.., 'SELF':[..], 'AI Rating':[..], 'BH':[..]}}."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Assessment Report"]
    cand = OrderedDict()
    cur = None
    name_of = {}
    for r in range(2, ws.max_row + 1):
        emp = ws.cell(row=r, column=2).value
        rev = ws.cell(row=r, column=9).value
        nm = ws.cell(row=r, column=5).value
        if emp:
            cur = emp
        if nm and cur:
            name_of[cur] = str(nm).strip()
        if not rev:
            continue
        scores = [_num(ws.cell(row=r, column=c).value) for c in COMP_COLS]
        cand.setdefault(cur, {})[str(rev).strip()] = scores
    for emp in cand:
        cand[emp]["_name"] = name_of.get(emp, emp)
    return cand


def load_ai_from_db():
    """Return {normalized_name: [10 competency averages]} from the live DB.

    Reads each completed session's report.competency_summary. Requires the
    project's database.py + DATABASE_URL. Read-only.
    """
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
    from database import list_sessions, get_session  # noqa: E402

    out = {}
    for s in list_sessions():
        if s.get("status") != "completed":
            continue
        full = get_session(s["session_id"])
        if not full:
            continue
        report = full.get("report") or {}
        cs = report.get("competency_summary") or {}
        if not cs:
            continue
        row = [cs.get(name) for name in COMP_NAMES]
        nm = (s.get("candidate_name") or "").strip().lower()
        if nm:
            out[nm] = row
    return out


def stats(pairs):
    if not pairs:
        return None
    a = [p[0] for p in pairs]
    b = [p[1] for p in pairs]
    n = len(pairs)
    bias = sum(x - y for x, y in zip(a, b)) / n
    mae = sum(abs(x - y) for x, y in zip(a, b)) / n
    ma, mb = statistics.mean(a), statistics.mean(b)
    sa, sb = statistics.pstdev(a), statistics.pstdev(b)
    cov = sum((x - ma) * (y - mb) for x, y in zip(a, b)) / n
    corr = cov / (sa * sb) if sa > 0 and sb > 0 else float("nan")
    w1 = sum(1 for x, y in zip(a, b) if abs(x - y) <= 1) / n
    w2 = sum(1 for x, y in zip(a, b) if abs(x - y) <= 2) / n
    return dict(n=n, bias=round(bias, 3), mae=round(mae, 3), corr=round(corr, 3),
                within1=round(w1 * 100, 1), within2=round(w2 * 100, 1),
                mean_ai=round(ma, 2), mean_bh=round(mb, 2))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--sheet-only", action="store_true",
                    help="Use the sheet's 'AI Rating' rows instead of the live DB")
    args = ap.parse_args()

    cand = load_sheet(args.xlsx)
    db_ai = None
    if not args.sheet_only:
        try:
            db_ai = load_ai_from_db()
            print(f"Loaded AI scores for {len(db_ai)} completed sessions from DB\n")
        except Exception as exc:
            print(f"[warn] DB load failed ({exc}); falling back to --sheet-only\n")
            db_ai = None

    pair_all = []
    per_comp = defaultdict(list)
    cand_rows = []
    matched = 0
    for emp, d in cand.items():
        bh = d.get("BH")
        if not bh:
            continue
        if db_ai is not None:
            ai = db_ai.get(d["_name"].lower())
        else:
            ai = d.get("AI Rating")
        if not ai:
            continue
        matched += 1
        pa, pb = [], []
        for i in range(10):
            if ai[i] is not None and bh[i] is not None and ai[i] >= 1:
                pair_all.append((ai[i], bh[i]))
                per_comp[COMP_NAMES[i]].append((ai[i], bh[i]))
                pa.append(ai[i]); pb.append(bh[i])
        if len(pa) >= 5:
            cand_rows.append((abs(statistics.mean(pa) - statistics.mean(pb)),
                              d["_name"], round(statistics.mean(pa), 2),
                              round(statistics.mean(pb), 2)))

    print(f"Matched candidates: {matched}\n")
    print("OVERALL AI vs BH:", json.dumps(stats(pair_all), indent=0))
    print("\nPER-COMPETENCY (bias = AI - BH):")
    for name in COMP_NAMES:
        st = stats(per_comp[name])
        if st:
            print(f"  {name[:34]:34s} n={st['n']:3d} bias={st['bias']:+.2f} "
                  f"mae={st['mae']:.2f} corr={st['corr']:.2f}")
    print("\nTOP OUTLIERS (|avgAI - avgBH|):")
    for g, nm, a, b in sorted(cand_rows, reverse=True)[:15]:
        print(f"  {nm[:24]:24s} avgAI={a:4.1f} avgBH={b:4.1f} gap={a-b:+.1f}")

    # Success-target check (intelligent interpreter, not BH-clone)
    o = stats(pair_all)
    if o:
        print("\nSUCCESS-TARGET CHECK:")
        print(f"  within ±1 ≥ 75%   : {o['within1']:.1f}%  {'PASS' if o['within1'] >= 75 else 'below'}")
        print(f"  within ±2 ≥ 90%   : {o['within2']:.1f}%  {'PASS' if o['within2'] >= 90 else 'below'}")
        print(f"  |bias| ≤ 0.3       : {abs(o['bias']):.2f}  {'PASS' if abs(o['bias']) <= 0.3 else 'below'}")
        print(f"  correlation ≥ 0.6  : {o['corr']:.2f}  {'PASS' if o['corr'] >= 0.6 else 'below'}")
        n_outliers = sum(1 for a, b in pair_all if abs(a - b) > 2)
        print(f"  >2-pt outliers     : {n_outliers} ({n_outliers/o['n']*100:.1f}% of {o['n']})")


if __name__ == "__main__":
    main()
